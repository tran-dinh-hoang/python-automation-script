import os
import openpyxl
from openpyxl import Workbook

# Function to extract hyperlinks from an Excel file (.xlsx)
def extract_hyperlinks_from_xlsx(xlsx_file):
    workbook = openpyxl.load_workbook(xlsx_file)
    hyperlinks = []

    # Iterate through all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        # Iterate through all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink:  # Check if the cell contains a hyperlink
                    title = cell.value  # Use the cell's value as the title
                    hyperlink = cell.hyperlink.target  # Get the hyperlink URL
                    hyperlinks.append((str(title).strip(), hyperlink))

    return hyperlinks

# Function to write hyperlinks to an Excel file in the output folder
def write_to_excel(file_name, data, output_folder):
    workbook = Workbook()
    sheet = workbook.active

    # First row: file name
    sheet["A1"] = "File Name"
    sheet["B1"] = file_name

    # Second row: headers
    sheet["A2"] = "Title"
    sheet["B2"] = "Hyperlink"

    # From the third row: data (title and hyperlink)
    for idx, (title, hyperlink) in enumerate(data, start=3):
        sheet[f"A{idx}"] = title
        sheet[f"B{idx}"] = hyperlink

    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)

    # Save the Excel file in the output folder
    output_file_name = os.path.basename(file_name).replace(".xlsx", "_hyperlinks.xlsx")
    output_file_path = os.path.join(output_folder, output_file_name)
    workbook.save(output_file_path)

    print(f"Data written to {output_file_path}")

# Main function
def main():
    # Folder where the Excel file is located
    input_folder = "input"
    
    # Folder where the Excel file with hyperlinks will be saved
    output_folder = "output"

    # List all files in the folder and select the first Excel file
    xlsx_files = [f for f in os.listdir(input_folder) if f.endswith(".xlsx")]

    if xlsx_files:
        xlsx_file_path = os.path.join(input_folder, xlsx_files[0])  # Get the first Excel file in the folder
        
        # Extract hyperlinks
        hyperlinks = extract_hyperlinks_from_xlsx(xlsx_file_path)

        # Write data to Excel
        if hyperlinks:
            write_to_excel(xlsx_file_path, hyperlinks, output_folder)
        else:
            print("No hyperlinks found in the Excel file.")
    else:
        print("No Excel (.xlsx) files found in the input folder.")

# Run the script
if __name__ == "__main__":
    main()
